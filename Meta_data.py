from bs4 import BeautifulSoup
import lxml
import requests
import time
import random
import openpyxl
import xlsxwriter

category_dict = {}                #set for storage of categories
workbook = openpyxl.load_workbook("Category_info.xlsx") 
# credits for user agent rotation method to https://www.scrapehero.com/how-to-fake-and-rotate-user-agents-using-python-3/
user_agent_list = [
   #Chrome
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
    'Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
    #Firefox
    'Mozilla/4.0 (compatible; MSIE 9.0; Windows NT 6.1)',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0)',
    'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.2; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/5.0)',
    'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0)',
    'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0)',
    'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; .NET CLR 2.0.50727; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729)'
]
sheet_obj = workbook.active 

workbook2 = xlsxwriter.Workbook("Meta_datainfo.xlsx")
sheet2 = workbook2.add_worksheet()
row=0
col=0
    
for j in range(1,2):
    URL = sheet_obj.cell(row = j, column = 2) 
    for i in range(1,11):    
        time.sleep(4)                                          #add delay to request
        
        user_agent = random.choice(user_agent_list)
        page = requests.get(URL.value,headers = {'User-Agent': user_agent})
        
        soup = BeautifulSoup(page.content,'lxml')           #parsing the html doc
        lib_url = soup.select('body > div[id = "page"] > div[id = "maincontent"] > div[class = "im"] > a[href ^= "/artifact/"]')  #structure of the website determined the tags to be used
        
        for temp in lib_url:
            new_url = "https://mvnrepository.com/"+temp['href']
            user_agent = random.choice(user_agent_list)
            time.sleep(3)                                          #add delay to request
            newpage = requests.get(new_url,headers = {'User-Agent': user_agent})
            newsoup = BeautifulSoup(newpage.content,'lxml')
            metainfo_heading = newsoup.select('body > div[id = "page"] > div[id = "maincontent"] > table[class = "grid"] > tr > th')  #structure of the website determined the tags to be used
            metainfo_value = newsoup.select('body > div[id = "page"] > div[id = "maincontent"] > table[class = "grid"] > tr > td')  #structure of the website determined the tags to be used
            
            for (head,val) in zip(metainfo_heading,metainfo_value):
                sheet2.write(row,col,head.text)
                sheet2.write(row,col+1,val.text)
                row += 1 
                
            
workbook.close()
workbook2.close()
