from bs4 import BeautifulSoup
import lxml
import requests
import time
import random
import openpyxl
import wget
import os
workbook = openpyxl.load_workbook("Category_info.xlsx") 

# credits for user agent rotation method to https://www.scrapehero.com/how-to-fake-and-rotate-user-agents-using-python-3/
user_agent_list = [
   #Chrome
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
	#Firefox
	'Mozilla/4.0 (compatible; MSIE 9.0; Windows NT 6.1)',
	'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (Windows NT 6.2; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0)',
	'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0)',
	'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; .NET CLR 2.0.50727; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729)'
]
sheet_obj = workbook.active 
	
for j in range(1,151):
	URL = sheet_obj.cell(row = j, column = 3)		#selecting category from Category_Info.xlsx
	cat_name = sheet_obj.cell(row = j,column = 1) 
	for i in range(1,11):    
		time.sleep(2)                                          #add delay to request
		url = URL.value + '?p='
		url = url + str(i)										#Url obtained for going to ith page in the category
		user_agent = random.choice(user_agent_list)
		page = requests.get(url,headers = {'User-Agent': user_agent})
		
		soup = BeautifulSoup(page.content,'lxml')           #parsing the html doc
		lib_url = soup.select('body > div[id = "page"] > div[id = "maincontent"] > div[class = "im"] > a[href ^= "/artifact/"]') #gets the set of urls for each library on the ith page of the jth category
		for temp in lib_url:
			new_url = "https://mvnrepository.com/"+temp['href']				#url of a library (say X) in the url set
			user_agent = random.choice(user_agent_list)
			time.sleep(2)                                          
			newpage = requests.get(new_url,headers = {'User-Agent': user_agent})
			newsoup = BeautifulSoup(newpage.content,'lxml')
			
			down_url = newsoup.select_one('table[class = "grid versions"] > tbody > tr > td > a[href]')	# url of the latest version of the jar library
			
			page_url = new_url+'/'+down_url.text		#url of the page from where download will be triggered
			lib_name = newsoup.select_one('h2[class="im-title"] > a[href]')		#name of the library being downloaded
			
			user_agent = random.choice(user_agent_list)
			time.sleep(2)                                          
			newpage2 = requests.get(page_url,headers = {'User-Agent': user_agent})
			newsoup2 = BeautifulSoup(newpage2.content,'lxml')		
			
			downlist = newsoup2.select('table[class = "grid"] a[href ^= "http://central.maven.org/maven2/"]')	#the page of the website features a view all button which is the true url containing list of files for download
			for temp in downlist:
				if temp.text == "View All":
					newpage3 = requests.get(temp['href'],headers = {'User-Agent': user_agent})
					newsoup3 = BeautifulSoup(newpage3.content,'lxml')
					finalurl = newsoup3.select('body a[href $= ".jar"]')			#obtaining the actual url for downloading the jar file
					path = "lib/"+cat_name.value+"/"+lib_name.text					#path for storing downloaded file
					
					try:
						os.makedirs(path)
						for down in finalurl:
							getit = temp['href']+'/'+down['href']			#downloading the file
							file = wget.download(getit,path)
					except:
						print("file already downloaded for: "+lib_name.text)	#if already downloaded then this message is displayed
						pass
workbook.close()