from bs4 import BeautifulSoup
import lxml
import requests
import time
import random
import openpyxl

workbook = openpyxl.load_workbook("Category_info.xlsx") 
wb2 = openpyxl.Workbook()
# credits for user agent rotation method to https://www.scrapehero.com/how-to-fake-and-rotate-user-agents-using-python-3/
user_agent_list = [
   #Chrome
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
	'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',
	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
	#Firefox
	'Mozilla/4.0 (compatible; MSIE 9.0; Windows NT 6.1)',
	'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (Windows NT 6.2; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)',
	'Mozilla/5.0 (Windows NT 6.1; Win64; x64; Trident/7.0; rv:11.0) like Gecko',
	'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0)',
	'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0)',
	'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; .NET CLR 2.0.50727; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729)'
]
sheet_obj = workbook.active 
sheet_obj2 = wb2.active
row = 1

for j in range(1,151):
	URL = sheet_obj.cell(row = j, column = 3) 
	for i in range(1,11):    
		time.sleep(2)                                          #add delay to request
		url = URL.value + '?p='
		url = url + str(i)
		user_agent = random.choice(user_agent_list)
		page = requests.get(url,headers = {'User-Agent': user_agent})
		
		soup = BeautifulSoup(page.content,'lxml')           #parsing the html doc
		lib_url = soup.select('body div[id = "page"] > div[id = "maincontent"] > div[class = "im"] > a[href ^= "/artifact/"]')  #structure of the website determined the tags to be used
		
		for temp in lib_url:
			new_url = "https://mvnrepository.com/"+temp['href']
			user_agent = random.choice(user_agent_list)
			time.sleep(2)
			newpage = requests.get(new_url,headers = {'User-Agent': user_agent})
			newsoup = BeautifulSoup(newpage.content,'lxml')
			lib_name = newsoup.select('body div[id = "page"] > div[id = "maincontent"] div[class = "im"] > div[class = "im-header"] > h2[class = "im-title"] > a[href]') #library name stored 
			metainfo_value = newsoup.select('body div[id = "page"] div[id = "maincontent"] > table[class = "grid"] > tr > td')  #this item stores meta data values
			
			items = 1
			col = 2
			col2 = 2 
			
			cell = sheet_obj2.cell(row=row,column=1)
			cell.value = lib_name[0].text			#storing name of library in xlsx
			print(cell.value)
			
			for valcount in metainfo_value:			#getting count of available meta data fields
				items = items + 1

			if items == 4:							#for some libraries usage field is not available hence substituting 0 for usage in such cases
				for val in metainfo_value:
					cell = sheet_obj2.cell(row=row,column=col2)
					cell.value = val.text			#structure of html table is slightly different hence the if else conditions to handle the capture of data
					col2 = col2 + 1
					print(cell.value)
					wb2.save("Meta_data.xlsx")					

				cell = sheet_obj2.cell(row=row,column=5)
				cell.value = 0
				row = row + 1

			elif items == 5:							#usage of library is given for this case

				for val in metainfo_value:
					cell = sheet_obj2.cell(row=row,column=col)
					if col == 5:
						try:
							numval = val.text.split()[:-1]
							numval = int(numval[0].replace(',',''))			#converting usage into integer value 
							cell.value = numval
						except:
							cell.value = 0
					else:
						cell.value = val.text								#for normal text data
					col = col + 1
					print(cell.value)
					wb2.save("Meta_data.xlsx")					
					if col == 6:
						col = 2
						row = row + 1

wb2.close()
workbook.close()